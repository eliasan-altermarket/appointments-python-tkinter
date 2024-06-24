# ############################ ############################
# Avgeros Kostas, 166187     
# avgerosk@gmail.com    
# Appointment Project 2023-2024 Hellenic Open University
# ############################ ############################
# https://www.geeksforgeeks.org/python-longest-string-in-list/
# ############################ ############################

import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from tkinter import filedialog, messagebox, Tk
from datetime import datetime


class ModuleExcel():
    '''A class used to represent Excel'''
    def __init__(self, date):
        '''
        Initialize Workbook and activates sheet
        '''
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        # self.todays_date = datetime.today().strftime('%Y-%m-%d')
        self.todays_date = date
        self.max_email_length = 0  # Initialize max email length
    
    def path_to_save(self):
        '''Returns user defined path to save the file '''
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                filetypes= [("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile= self.todays_date)
        return file_path

    def add_styles(self):
        '''Add font, alignment, border, color in cells '''
        red_font = Font(name='Comic Sans', color='00FF0000', bold=True, size=10)
        alignment = Alignment(horizontal='center',
                vertical='center',
                shrink_to_fit=False,
                indent=0)
        pink = "00FF00FF"
        border = Side(border_style="thin", color=pink)
        # Enumerate
        for cell in self.sheet["1:1"]:
            cell.font = red_font
        for row in self.sheet.iter_cols():
            for cell in row:
                cell.alignment = alignment 
                cell.border = Border(top=border, left=border, right=border, bottom=border)

    def add_heading(self):
        '''Add headings in excel file as Template'''
        heading = ["Name", "Last name", "Slot", "AppointmentID", "Email"]
        self.sheet.append(heading)
        
    def add_data(self, email_q, slot_q, AppointmentID_q, firstname_q, lastname_q):
        '''Appends data params in to excel row by row'''
        self.sheet.append([ firstname_q, lastname_q, slot_q, AppointmentID_q, email_q ])
        # Longest String in list, using max() + key, res = max(test_list, key = len)
        self.max_email_length = max(self.max_email_length, len(email_q))
        print("Maximum rows after append:", self.sheet.max_row)

    def del_previous_data(self, file_path):
        '''Deletes previous data escept Headers'''
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active
        self.sheet.delete_rows(2, self.sheet.max_row-1)
        print("\tDELETES rows except excel Headers")

    def set_header_dimensions(self): 
        '''Set dimensions using dictionary method only for Header, usefull if no data exists'''
        extend_len = 5
        columns = {
            'B': str('lastname_q'),
            'D': str('AppointmentID_q')
        }
        for col, value in columns.items():
            self.sheet.column_dimensions[col].width = len(value) + extend_len

    def set_dimensions(self, slot_q, AppointmentID_q, firstname_q, lastname_q): # without email_q
        '''Set dimensions using dictionary method'''
        extend_len = 5
        columns = {
            'A': str('firstname_q'), # TODO if last firstname is smaller
            'B': str('firstname_q'), # TODO if last lastname is smaller
            'C': slot_q,
            'D': str('AppointmentID_q'),
            'E': ' ' * self.max_email_length
        }
        for col, value in columns.items():
            self.sheet.column_dimensions[col].width = len(value) + extend_len
        
    def save(self, file_path):
        '''A function to save file'''
        try:
            self.workbook.save(file_path)
            print(f'File saved as {file_path}')
        except PermissionError as e:
            popup_error = Tk()
            popup_error.withdraw()  # Hide the root window
            messagebox.showerror("Error", "File is already open. Please close the file and try again.")
            popup_error.destroy()  # Destroy the root window after the message box is closed
            print("File is already open. Please close the file and try again.", e) # prints twice cause twice it writes
